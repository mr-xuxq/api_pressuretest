# -*-coding:utf-8-*-
"""
File Name:do_excel.py
Program IDE:PyCharm
Create File Time:2022/7/4 3:18 PM
File Create By Author:xuxiaoqi
"""
from openpyxl import load_workbook
from conf import read_path

class DoExcel:
    def __init__(self, file_path, sheet_name):
        self.file_path=file_path
        self.sheet_name=sheet_name

    def do_excel(self):
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        test_excel_data = []

        for i in range(2, sheet.max_row+1):
            test_dict = {}
            test_dict["case_id"] = sheet.cell(i, 1).value
            test_dict["case_name"] = sheet.cell(i, 2).value
            test_dict["method"] = sheet.cell(i, 3).value
            test_dict["url"] = sheet.cell(i, 4).value
            test_dict["param"] = sheet.cell(i, 5).value
            test_dict["expected_1"] = sheet.cell(i, 6).value
            test_dict["expected_2"] = sheet.cell(i, 7).value
            test_dict["concurrency"] = sheet.cell(i, 11).value
            test_dict["cycles"] = sheet.cell(i, 12).value
            test_excel_data.append(test_dict)
        return test_excel_data

    def write_data(self, r, c, value):
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        sheet.cell(r, c).value = value
        wb.save(self.file_path)

if __name__ == "__main__":
    file_path = read_path.test_data_path
    sheet_name = 'business_collect_global'
    test_excel_data=DoExcel(file_path, sheet_name).do_excel()
    for i in range(len(test_excel_data)):
        print(test_excel_data[i])
    #DoExcel(file_path,sheet_name).write_data(len(test_excel_data)+2,1,len(test_excel_data)+1)











